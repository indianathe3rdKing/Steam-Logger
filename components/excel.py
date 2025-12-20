# Import necessary libraries for date/time and Excel operations
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from data_fetch import normalized_rows_aspen, normalized_rows_fresenius
from flask import Flask, send_file
import io
from appwrite_db import storage, bucket_Id, aspen_file_Id, fresenius_file_Id
from appwrite.input_file import InputFile
# Name of the Excel files to save data to
AspenFileName = "steam_data_aspen.xlsx"
FreseniusFileName = "steam_data_fresenius.xlsx"

app = Flask(__name__)

# Column headers for the Excel spreadsheet
HEADERS_ASPEN = [
    "Date", "Time", "Meter 1", "Bypass", "Meter Blue", "Meter Red", "Steam Flow Meter", "Aspen"
]
HEADERS_FRESENIUS = [
    "Date", "Time", "Meter FK", "Make Up", "Meter SH", "HFO", "Steam Flow Meter 1", "Steam Flow Meter 2"
]

# Get current date and time
today = datetime.datetime.now()


# Loads or creates an Excel workbook and creates a new sheet for the current month
def get_workbook(headers, file_Id):

    result = storage.get_file_download(
        bucket_id=bucket_Id,
        file_id=file_Id,
    )

    excel_data = io.BytesIO(result)

    # If file exists, open it and add a new sheet named after the current month
    wb = load_workbook(excel_data)
    ws = wb.create_sheet(title=today.strftime("%B"))
    ws.append(headers)
    # Style the header row
    for col in ws['1:1']:
        col.font = Font(size=12, bold=True)
        col.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col.column_letter].width = 15
        ws.column_dimensions["G"].width = 20
        # thick = Side(style="thick", color="000000")
        # ws[col.column_letter+"1"].border = Border(
        #     left=thick, right=thick, top=thick, bottom=thick)
        col.fill = PatternFill("solid", start_color="A8BBA3")

    return wb, ws
#     # If file doesn't exist, create a new workbook with headers
#     wb = Workbook()
#     ws = wb.active
#     ws.title = x.strftime("%B")
#     ws.append(HEADERS)
#     # Style the header row
#     for col in ws['1:1']:
#         col.font = Font(size=12, bold=True)
#         col.alignment = Alignment(horizontal="center")
#         ws.column_dimensions[col.column_letter].width = 15
#         ws.column_dimensions["G"].width = 20
#         # thin = Side(style="thin", color="000000")
#         # ws[col.column_letter+"1"].border = Border(
#         #     left=thin, right=thin, top=thin, bottom=thin)
#         col.fill = PatternFill("solid", start_color="A8BBA3")
# return wb, ws


# Adds meter reading rows to the Excel sheet and saves the file


def append_rows(rows, workbook, last_column):

    wb, ws = workbook

    # Iterate through each row and add it to the worksheet
    for row in rows:
        ws.append(row)
        # print(f"Appended row: {row}")
        # style the cells
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", start_color="EBF4DD")

    # Get difference between last row and second last row for meter readings
    last_row = ws.max_row
    summary_row = last_row + 1
    for col in range(3, last_column + 1):
        first_value = ws.cell(row=2, column=col).value
        last_value = ws.cell(row=last_row, column=col).value

        try:
            first_value = float(first_value)
            last_value = float(last_value)
            difference = last_value - first_value
        except (TypeError, ValueError):
            difference = None

        cell = ws.cell(row=summary_row, column=col)
        cell.value = difference
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", start_color="FFA239")
        cell.font = Font(bold=True)

    # Save the updated workbook to disk
    updated_workbook = io.BytesIO()
    wb.save(updated_workbook)
    updated_workbook.seek(0)

    return updated_workbook


# Execute: Normalize the fetched data and write to Excel files

def update_aspen_excel_file():
    try:
        workbook_aspen = get_workbook(HEADERS_ASPEN, aspen_file_Id)
        File = append_rows(normalized_rows_aspen, workbook_aspen, 8)
        storage.delete_file(
            bucket_id=bucket_Id,
            file_id=aspen_file_Id,
        )
        storage.create_file(
            bucket_id=bucket_Id,
            file_id=aspen_file_Id,
            file=InputFile.from_bytes(File.read(), AspenFileName)
        )
        print("Aspen Excel file updated successfully!")
    except Exception as e:
        print(f"Error updating Aspen Excel file: {e}")


def update_fresenius_excel_file():
    try:
        workbook_fresenius = get_workbook(HEADERS_FRESENIUS, fresenius_file_Id)
        File = append_rows(normalized_rows_fresenius, workbook_fresenius, 8)
        storage.delete_file(
            bucket_id=bucket_Id,
            file_id=fresenius_file_Id,
        )
        storage.create_file(
            bucket_id=bucket_Id,
            file_id=fresenius_file_Id,
            file=InputFile.from_bytes(File.read(), FreseniusFileName)
        )
        print("Fresenius Excel file updated successfully!", File)
    except Exception as e:
        print(f"Error updating Fresenius Excel file: {e}")


if __name__ == "__main__":
    print("Starting Excel update...")
    update_fresenius_excel_file()
    update_aspen_excel_file()
    print("Excel update script completed.")
